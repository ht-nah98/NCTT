"""Nạp bảng DNA âm thanh (v2) → parquet.

NGUỒN: <N>/00_input/raw/audio_dna.xlsx — bản trích đặc trưng nhiều bản nhạc,
mỗi dòng một TRACK (một video có thể chứa nhiều track, cắt theo chapter).

KHÁC normalize_audio.py: file kia đọc YAML mỗi bản một file (v1, n=5).
File này đọc bảng nhiều dòng (v2, n=307). Hai nguồn ĐỘC LẬP, không đè nhau.

CỘT BỊ BỎ: 'prompt day du' (sheet prompt Suno) — người dùng xác nhận
dữ liệu thô cột này đang SAI, không phân tích cho tới khi có bản sửa.
"""
import sys
import pandas as pd

sys.path.insert(0, str(__import__("pathlib").Path(__file__).resolve().parents[1]))
from _common import niche_paths

N, P, _ = niche_paths()

SRC = N / "00_input/raw/audio_dna.xlsx"
SHEET_TRACKS = "307 track"      # tên sheet bản v2
SHEET_CORPUS = "thong ke corpus"

# Cột đúng ra phải là số. Excel hay để lẫn chuỗi rỗng / 'nan'.
NUMERIC = ["phut", "bpm", "nhip", "tempo_cv", "swing_phase", "dao_phach", "do_tre_ms",
           "four_on_floor", "tong_tin_cay", "hop_am_moi_o_nhip", "so_hop_am_rieng",
           "lufs", "plr_db", "lra", "stereo_width", "stem_vocals", "stem_bass",
           "stem_drums", "stem_guitar", "stem_piano", "tach_stem_dB", "quang_semitone",
           "buoc_lien", "not_moi_giay", "vibrato_hz", "hnr_db", "jitter", "lech_cent",
           "bam_luoi_semitone", "tuong_quan_LR"]
BOOL = ["nhip_nhap_nhang", "tong_khop_CNN", "tach_stem_yeu", "co_giong_hat"]


def read_sheet(path, name):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if name not in wb.sheetnames:
        # tên sheet có thể đổi theo số lượng track ("307 track" → "512 track")
        cands = [s for s in wb.sheetnames if "track" in s.lower()]
        if not cands:
            raise SystemExit(f"Không thấy sheet track trong {path}: {wb.sheetnames}")
        name = cands[0]
    rows = list(wb[name].iter_rows(values_only=True))
    return pd.DataFrame(rows[1:], columns=rows[0])


def main():
    if not SRC.exists():
        print(f"⏭  Bỏ qua: chưa có {SRC}")
        return
    T = read_sheet(SRC, SHEET_TRACKS)
    T = T[T.track_id.notna()].copy()

    for c in NUMERIC:
        if c in T:
            T[c] = pd.to_numeric(T[c], errors="coerce")
    for c in BOOL:
        if c in T:
            T[c] = T[c].map({True: True, False: False, "True": True, "False": False})

    # video_id là khoá nối sang video_master. Không có thì track vô dụng.
    before = len(T)
    T = T[T.video_id.notna()]
    if len(T) < before:
        print(f"   ⚠ bỏ {before-len(T)} track thiếu video_id")

    T.to_parquet(P / "audio_dna.parquet", index=False)

    # bảng thống kê corpus — giữ nguyên, dùng làm mốc so sánh ngách khác
    try:
        C = read_sheet(SRC, SHEET_CORPUS)
        C.to_parquet(P / "audio_dna_corpus.parquet", index=False)
    except SystemExit:
        C = None

    print(f"✅ {P/'audio_dna.parquet'}  {len(T)} track / {T.video_id.nunique()} video")
    if C is not None:
        print(f"✅ {P/'audio_dna_corpus.parquet'}  {len(C)} chỉ số corpus")


if __name__ == "__main__":
    main()
